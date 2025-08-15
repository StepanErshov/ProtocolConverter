import pandas as pd
import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Protection, Alignment
from openpyxl.utils import get_column_letter
from copy import copy
import os
from pathlib import Path
from io import BytesIO
from datetime import datetime
from xlsx2dbcForRelease import ExcelToDBCConverter
# Формирование папок
import create_directory
# Замер времени выполнения блоков
import time
# Многопоточность
import concurrent.futures
import zipfile

def set_page_title():
    st.title("🛎️ Release Convertor")

def get_uploaded_file():
    uploaded_file = st.file_uploader("Upload Domain Excel Matrix", type=["xlsx"], accept_multiple_files=True)
    if uploaded_file:
        return uploaded_file

def process_matrix_sheet(df, domain_matrix, ecu_matrices, ecu_col_index, progress_bar=None):
    process_matrix_sheet_time_start = time.time()
    
    # Prepare data structures
    ecu_rows_to_copy = {}
    for ecu_name, ecu_index in ecu_col_index.items():
        rows_to_copy = df[df[ecu_name].notna()].index.tolist()
        ecu_rows_to_copy[ecu_name] = [-1] + rows_to_copy
    
    print("Obtained ecu_rows_to_copy")
    
    domain_ws = domain_matrix['Matrix']
    max_column = domain_ws.max_column  # Pre-calculate before threading
    
    # Initialize progress
    total_tasks = len(ecu_rows_to_copy)
    completed_tasks = 0
    
    if progress_bar:
        progress_bar.progress(0, text="Processing matrix sheets...")
    
    # We'll use a thread pool executor for better parallel execution
    with concurrent.futures.ThreadPoolExecutor() as executor:
        # Create futures for each ECU processing task
        futures = {
            executor.submit(process_single_ecu, 
                           ecu_name, 
                           row_list_to_copy, 
                           domain_ws, 
                           ecu_matrices[ecu_name]['Matrix'], 
                           max_column): ecu_name
            for ecu_name, row_list_to_copy in ecu_rows_to_copy.items()
        }
        
        # Wait for all tasks to complete
        for future in concurrent.futures.as_completed(futures):
            ecu_name = futures[future]
            try:
                future.result()  # This will re-raise any exceptions from the thread
                completed_tasks += 1
                if progress_bar:
                    progress = int(completed_tasks / total_tasks * 100)
                    progress_bar.progress(progress, text=f"Processed {completed_tasks}/{total_tasks} matrix sheets")
                print(f'{ecu_name} matrix-sheet processed')
            except Exception as e:
                print(f'Error processing {ecu_name}: {str(e)}')

    process_matrix_sheet_time_end = time.time() - process_matrix_sheet_time_start
    print("process_matrix_sheet_time_end", process_matrix_sheet_time_end)

    return ecu_matrices

def process_single_ecu(ecu_name, row_list_to_copy, domain_ws, ecu_ws_matrix, max_column):
    """Process a single ECU's matrix sheet (thread-safe operations)"""
    # Process the rows and columns
    row_to_paste = 1
    for row_idx in row_list_to_copy:
        for col_idx in range(1, max_column + 1):
            # Get source cell
            domain_matrix_cell = domain_ws.cell(row=row_idx+2, column=col_idx)
            ecu_matrix_cell = ecu_ws_matrix.cell(row=row_to_paste, column=col_idx)
            
            # Copy value
            ecu_matrix_cell.value = domain_matrix_cell.value
            
            # Copy formatting if it exists
            if domain_matrix_cell.has_style:
                ecu_matrix_cell.font = copy(domain_matrix_cell.font)
                ecu_matrix_cell.border = copy(domain_matrix_cell.border)
                ecu_matrix_cell.fill = copy(domain_matrix_cell.fill)
                ecu_matrix_cell.number_format = copy(domain_matrix_cell.number_format)
                ecu_matrix_cell.protection = copy(domain_matrix_cell.protection)
                ecu_matrix_cell.alignment = copy(domain_matrix_cell.alignment)
        row_to_paste += 1
    
    # Group rows by messages
    row_idx = 2
    while row_idx <= ecu_ws_matrix.max_row:
        a_val = ecu_ws_matrix.cell(row=row_idx, column=1).value
        ecu_ws_matrix.row_dimensions[row_idx].height = 20
        if a_val:
            start_row = row_idx + 1
            end_row = start_row
            while end_row <= ecu_ws_matrix.max_row:
                i_val = ecu_ws_matrix.cell(row=end_row, column=9).value
                if i_val:
                    end_row += 1
                else:
                    break
            if end_row > start_row:
                for r in range(start_row, end_row):
                    ecu_ws_matrix.row_dimensions[r].outlineLevel = 1
                    ecu_ws_matrix.row_dimensions[r].hidden = True
                ecu_ws_matrix.row_dimensions[row_idx].collapsed = True
            row_idx = end_row
        else:
            row_idx += 1

def process_history_sheet(df, domain_matrix, ecu_matrices, ecu_col_index, progress_bar=None):
    process_history_sheet_time_start = time.time()

    if "History" not in domain_matrix.sheetnames:
        return False

    # Prepare the DataFrame
    df.columns = df.iloc[0]
    df = df.drop(0).reset_index(drop=True)
    
    # Create a dictionary to store rows to copy for each ECU
    history_rows_to_copy = {}
    for ecu_name, ecu_index in ecu_col_index.items():
        mask = df["ECU\n节点"].notna() & df["ECU\n节点"].str.contains(ecu_name, case=True)
        rows_to_copy = df[mask].index.tolist()
        history_rows_to_copy[ecu_name] = [-1] + rows_to_copy
    
    # Get all needed data from domain_ws before threading
    domain_ws = domain_matrix['History']
    max_column = domain_ws.max_column
    
    # Pre-load all cell data we'll need (thread-safe preparation)
    cell_data = {}
    for row_list in history_rows_to_copy.values():
        for row_idx in row_list:
            actual_row = row_idx + 3
            if actual_row > 0:  # Skip the -1 case
                for col_idx in range(1, max_column + 1):
                    cell = domain_ws.cell(row=actual_row, column=col_idx)
                    cell_data[(actual_row, col_idx)] = {
                        'value': cell.value,
                        'font': copy(cell.font) if cell.has_style else None,
                        'border': copy(cell.border) if cell.has_style else None,
                        'fill': copy(cell.fill) if cell.has_style else None,
                        'number_format': copy(cell.number_format) if cell.has_style else None,
                        'protection': copy(cell.protection) if cell.has_style else None,
                        'alignment': copy(cell.alignment) if cell.has_style else None
                    }
    
    # Get header data (single read before threading)
    header_cell = domain_ws.cell(row=1, column=1)
    header_data = {
        'value': header_cell.value,
        'font': copy(header_cell.font) if header_cell.has_style else None,
        'border': copy(header_cell.border) if header_cell.has_style else None,
        'fill': copy(header_cell.fill) if header_cell.has_style else None,
        'number_format': copy(header_cell.number_format) if header_cell.has_style else None,
        'protection': copy(header_cell.protection) if header_cell.has_style else None,
        'alignment': copy(header_cell.alignment) if header_cell.has_style else None
    }

    def process_single_history_ecu(ecu_name, row_list_to_copy):
        """Process history sheet for a single ECU"""
        history_ws_matrix = ecu_matrices[ecu_name]['History']
        
        # Process header
        history_ws_matrix.merge_cells('A1:G1')
        ecu_header_cell = history_ws_matrix.cell(row=1, column=1)
        ecu_header_cell.value = header_data['value']
        if header_data['font']:
            ecu_header_cell.font = header_data['font']
            ecu_header_cell.border = header_data['border']
            ecu_header_cell.fill = header_data['fill']
            ecu_header_cell.number_format = header_data['number_format']
            ecu_header_cell.protection = header_data['protection']
            ecu_header_cell.alignment = header_data['alignment']
        
        # Process each row
        row_to_paste = 2
        for row_idx in row_list_to_copy:
            actual_row = row_idx + 3
            if actual_row > 0:  # Skip the -1 case
                for col_idx in range(1, max_column + 1):
                    data = cell_data[(actual_row, col_idx)]
                    ecu_cell = history_ws_matrix.cell(row=row_to_paste, column=col_idx)
                    ecu_cell.value = data['value']
                    if data['font']:
                        ecu_cell.font = data['font']
                        ecu_cell.border = data['border']
                        ecu_cell.fill = data['fill']
                        ecu_cell.number_format = data['number_format']
                        ecu_cell.protection = data['protection']
                        ecu_cell.alignment = data['alignment']
                
                history_ws_matrix.row_dimensions[row_to_paste].height = 15
                row_to_paste += 1

        return ecu_name

    # Initialize progress
    total_tasks = len(history_rows_to_copy)
    completed_tasks = 0
    
    if progress_bar:
        progress_bar.progress(0, text="Processing history sheets...")

    # Use ThreadPoolExecutor for parallel processing
    with concurrent.futures.ThreadPoolExecutor() as executor:
        # Submit all ECU processing tasks
        future_to_ecu = {
            executor.submit(
                process_single_history_ecu,
                ecu_name,
                row_list_to_copy
            ): ecu_name
            for ecu_name, row_list_to_copy in history_rows_to_copy.items()
        }
        # print("future_to_ecu: ", future_to_ecu)
        # Process results as they complete
        for future in concurrent.futures.as_completed(future_to_ecu):
            ecu_name = future_to_ecu[future]
            try:
                result = future.result()
                completed_tasks += 1
                if progress_bar:
                    progress = int(completed_tasks / total_tasks * 100)
                    progress_bar.progress(progress, text=f"Processed {completed_tasks}/{total_tasks} history sheets")
                print(f'{result} history-sheet processed')
            except Exception as e:
                print(f'Error processing {ecu_name}: {str(e)}')
    
    process_history_sheet_time_end = time.time() - process_history_sheet_time_start
    print(f"Processed history sheet in {process_history_sheet_time_end:.2f} seconds")

    return ecu_matrices

def identify_ecus(df):
    # Найти все ecu в предоставленной доменной матрице
    ecus = [
        col
        for col in df.columns
        if any(val in ["S", "R"] for val in df[col].dropna().unique())
        and col != "Unit\n单位"
    ]

    if not ecus:
        st.error(
            "🕭 No ECUs found in the file. Please check that the file contains columns with 'S' or 'R' values."
        )
        st.stop()

    return ecus

def get_ecu_version(df_history, ecu_matrices):
    get_ecu_version_time_start = time.time()
    
    ecu_versions = {ecu_name : None for ecu_name in ecu_matrices}
    ecu_versions_checkbox = list(ecu_versions.keys())
    version_column = "Revision\n版本"
    ecu_column = "ECU\n节点"

    for idx in reversed(df_history.index):
        row = df_history.loc[idx]
        
        # Skip if first column is empty
        if pd.isna(row[version_column]) or row[version_column] == '':
            continue
            
        # Get value from target column
        ecus_mentioned = row[ecu_column]
        ecus_mentioned = ecus_mentioned.split(',')
        if any(ecu in ecus_mentioned for ecu in ecu_versions_checkbox):
            for ecu in ecus_mentioned:
                if ecu_versions[ecu] is None:
                    ecu_versions[ecu] = row[version_column]
                    ecu_versions_checkbox.remove(ecu)
        if not ecu_versions_checkbox:
            break
    
    get_ecu_version_time_end = time.time() - get_ecu_version_time_start
    print("get_ecu_version_time_end", get_ecu_version_time_end)

    return ecu_versions

def get_domain_folder_name(ecu_base, domain_short):
    if ecu_base in ("SGW", "CGW", "ADCU"):
        domain_short = "SGW-CGW"
    domain_folder_name = {
        "BD" :      "04.01.01.Body CAN",
        "DG" :      "04.01.08 Diagnostic CAN",
        "CH" :      "04.01.05.Chassis CANFD",
        "PT" :      "04.01.02.Powertrain CAN",
        "ET" :      "04.01.04.Entertainment CANFD",
        "DZ" :      "04.01.06.Demilitary zone CANFD",
        "SGW-CGW" : "04.01.07.CGW,SGW,ADCU"
    }

    return domain_folder_name[domain_short]

def get_ecu_folder_name(domain_folder_name, ecu_base):
    ecu_folder_name = ""
    for sub_folder_name in create_directory.creator.HIERARCHI[domain_folder_name]:
        if ecu_base in sub_folder_name:
            ecu_folder_name = sub_folder_name
            break
    if ecu_folder_name:
        return ecu_folder_name
    else:
        print("Sub folder for ECU doesn't match ECU name.", )
        # st.error("Sub folder for ECU doesn't match ECU name.")

def get_ecu_matrix_template(uploaded_file, ecu_col_index):
    col_index_to_letter = {i: get_column_letter(i) for i in range(1, 51)}
    ecu_wb = load_workbook(uploaded_file)
    matrix_ws_with_data = ecu_wb["Matrix"]
    history_ws_with_data = ecu_wb["History"]
    ecu_wb.remove(matrix_ws_with_data)
    ecu_wb.remove(history_ws_with_data)
    ecu_wb.create_sheet("Matrix")
    ecu_wb.create_sheet("History")
    
    # Сохранить ширину столбцов
    for col_idx in col_index_to_letter:
        if col_idx in ecu_col_index.values():
            ecu_wb["Matrix"].column_dimensions[col_index_to_letter[col_idx+1]].width = 2
        else:
            col_dimension = matrix_ws_with_data.column_dimensions[col_index_to_letter[col_idx]].width
            ecu_wb["Matrix"].column_dimensions[col_index_to_letter[col_idx]].width = col_dimension
    
    return ecu_wb

def save_single_ecu(ecu_name, ecu_matrices, ecu_versions, domain_short, converter, progress_callback=None):
    """Helper function to save a single ECU matrix and convert to DBC"""
    try:
        # Get base ECU name
        ecu_base = ecu_name.split("_")[0] if '_' in ecu_name else ecu_name
        
        # Get folder paths
        domain_folder_name = get_domain_folder_name(ecu_base, domain_short)
        ecu_folder_name = get_ecu_folder_name(domain_folder_name, ecu_base)
        
        if not ecu_folder_name:
            st.warning(f"ECU folder name not found for {ecu_base}. Skipping.")
            return None
        
        # Prepare output paths
        date_str = datetime.now().strftime("%d%m%Y")
        output_ecu_filename = f"ATOM_CAN_Matrix_{ecu_versions[ecu_name]}_{date_str}_{ecu_name}"
        
        # Create full directory path
        full_dir_path = os.path.join(create_directory.creator.PATH_DOC, domain_folder_name, ecu_folder_name)
        
        # Create directory if it doesn't exist
        os.makedirs(full_dir_path, exist_ok=True)
        
        ecu_matrix_output_path = os.path.join(full_dir_path, f"{output_ecu_filename}.xlsx")
        
        # Save the Excel file
        save_time_start = time.time()
        ecu_matrices[ecu_name].save(ecu_matrix_output_path)
        save_time = time.time() - save_time_start
        
        # if progress_callback:
        #     progress_callback("saved")
        
        # Convert to DBC
        dbc_time_start = time.time()
        dbc_path = os.path.join(full_dir_path, f"{output_ecu_filename}.dbc")
        
        # Initialize converter with the saved XLSX file
        print("ecu_matrix_output_path: ", ecu_matrix_output_path, "\n", "ecu_name: ", ecu_name, "\n", "save_time: ", save_time, "\n", "dbc_path: ", dbc_path)
        dbc_converter = ExcelToDBCConverter(ecu_matrix_output_path)
        conversion_success = dbc_converter.convert(dbc_path)
        dbc_time = time.time() - dbc_time_start
        
        # if progress_callback:
        #     progress_callback("converted")
        
        if not conversion_success:
            return (ecu_name, save_time, None, ecu_matrix_output_path, "DBC conversion failed")
        
        return (ecu_name, save_time, dbc_time, ecu_matrix_output_path, dbc_path)
    except Exception as e:
        import traceback
        print(f"❌ Ошибка при обработке ECU {ecu_name}:")
        traceback.print_exc()
        return (ecu_name, None, None, None, f"Exception: {repr(e)}")

if __name__ == "__main__":
    set_page_title()
    uploaded_file = get_uploaded_file()
    for file in uploaded_file:
        if file:
            try:
                # Create progress container
                progress_container = st.empty()
                progress_bar = progress_container.progress(0)
                status_text = st.empty()
                
                total_start = time.time()
                status_text.text("Reading uploaded file...")
                print("File uploaded")
                df_matrix = pd.read_excel(file, sheet_name="Matrix")
                df_history = pd.read_excel(file, sheet_name="History")
                
                status_text.text("Identifying ECUs...")
                ecus = identify_ecus(df_matrix)
                domain_matrix = load_workbook(file)
                # Получить номер столбца для каждого ecu
                ecu_col_index = {ecu: df_matrix.columns.get_loc(ecu) for ecu in ecus}

                status_text.text("Creating ECU matrices...")
                print("Ecu matrices start creating")
                # Создать пустые excel матрицы для каждого ecu
                ecu_matrices = {}
                # Create ThreadPoolExecutor
                with concurrent.futures.ThreadPoolExecutor() as executor:
                    # Submit all ECU template loading tasks
                    future_to_ecu = {
                        executor.submit(get_ecu_matrix_template, file, ecu_col_index): ecu_name
                        for ecu_name in ecu_col_index
                    }
                    
                    # Process results as they complete
                    for future in concurrent.futures.as_completed(future_to_ecu):
                        ecu_name = future_to_ecu[future]
                        try:
                            ecu_matrices[ecu_name] = future.result()
                            progress = len(ecu_matrices) / len(ecu_col_index) * 100
                            progress_bar.progress(int(progress))
                            status_text.text(f"Initialized {len(ecu_matrices)}/{len(ecu_col_index)} ECU matrices")
                            print(f"Initialized matrix for {ecu_name}")
                        except Exception as e:
                            print(f"Error loading template for {ecu_name}: {str(e)}")
                print("Ecu matrices created")

                # Настроить лист "Matrix" под ecu
                status_text.text("Processing matrix sheets...")
                ecu_matrices = process_matrix_sheet(df_matrix, domain_matrix, ecu_matrices, ecu_col_index, progress_bar)
                
                status_text.text("Processing history sheets...")
                ecu_matrices = process_history_sheet(df_history, domain_matrix, ecu_matrices, ecu_col_index, progress_bar)
                if not ecu_matrices:
                    st.warning(f"🕱 'History' sheet not found in ECU {ecu_matrices}.")
                
                status_text.text("Getting ECU versions...")
                ecu_versions = get_ecu_version(df_history, ecu_matrices)
                
                if not create_directory:
                    st.warning("'create_directory' module is missing. Skipping file save.")
                
                domain_short = file.name.split('_')[3]
                
                # Сохранение файлов
                save_results = []
                total_ecus = len(ecu_col_index)
                
                # Progress tracking for saving
                # saved_count = 0
                # converted_count = 0
                
                def update_progress(action):
                    saved_count = 0
                    converted_count = 0
                    if action == "saved":
                        saved_count += 1
                    elif action == "converted":
                        converted_count += 1
                    
                    progress_text = f"Saving files... ({saved_count}/{total_ecus} saved)"
                    if saved_count == total_ecus:
                        progress_text = f"Converting to DBC... ({converted_count}/{total_ecus} converted)"
                    
                    status_text.text(progress_text)
                    progress = (saved_count + converted_count) / (total_ecus * 2) * 100
                    progress_bar.progress(int(progress))
                
                with concurrent.futures.ThreadPoolExecutor() as executor:
                    futures = {
                        executor.submit(
                            save_single_ecu,
                            ecu_name,
                            ecu_matrices,
                            ecu_versions,
                            domain_short,
                            ExcelToDBCConverter,
                            None
                        ): ecu_name for ecu_name in ecu_col_index
                    }

                    for future in concurrent.futures.as_completed(futures):
                        ecu_name = futures[future]
                        try:
                            result = future.result()
                            if result is None:
                                st.error(f"❌ {ecu_name}: не удалось создать матрицу")
                                save_results.append((ecu_name, None, None, None, "Matrix creation failed"))
                            else:
                                if len(result) != 5:
                                    st.error(f"❌ {ecu_name}: неверное количество данных: {len(result)}")
                                    save_results.append((ecu_name, None, None, None, "Invalid result format"))
                                else:
                                    save_results.append(result)
                        except Exception as e:
                            st.error(f"❌ Ошибка при обработке {ecu_name}: {str(e)}")
                            save_results.append((ecu_name, None, None, None, f"Exception: {str(e)}"))

                proccesed_time = time.time() - total_start
                print("proccesed_time", proccesed_time)

                progress_bar.empty()
                status_text.empty()

                st.success(f"Domain matrix ECU split completed, obtained {len(ecu_col_index)} ECUs with DBC files.")
                st.info(f"Time spent: {proccesed_time:.2f} seconds")
                st.info(f"XLSX files saved: {len([r for r in save_results if r[1] is not None])}")
                st.info(f"DBC files generated: {len([r for r in save_results if r[2] is not None])}")
                            
                import zipfile
                from io import BytesIO

                zip_buffer = BytesIO()
                with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                    for result in save_results:
                        if len(result) < 5:
                            continue
                        ecu_name, save_time, dbc_time, xlsx_path, dbc_path_or_error = result

                        if xlsx_path and isinstance(xlsx_path, str) and os.path.exists(xlsx_path):
                            try:
                                relative_path = os.path.relpath(xlsx_path, create_directory.creator.PATH_DOC)
                                zip_file.write(xlsx_path, arcname=relative_path)
                            except Exception as e:
                                print(f"❌ Не удалось добавить XLSX в ZIP: {e}")

                        if isinstance(dbc_path_or_error, str) and os.path.exists(dbc_path_or_error):
                            try:
                                relative_path = os.path.relpath(dbc_path_or_error, create_directory.creator.PATH_DOC)
                                zip_file.write(dbc_path_or_error, arcname=relative_path)
                            except Exception as e:
                                print(f"❌ Не удалось добавить DBC в ZIP: {e}")

                zip_buffer.seek(0)

                st.download_button(
                    label="📥 Скачать всё с папками (ZIP)",
                    data=zip_buffer,
                    file_name=f"Generated_CAN_Files_{ecu_name}.zip",
                    mime="application/zip",
                    disabled=zip_buffer.getbuffer().nbytes == 0
                )


                # Show download buttons for DBC files
                for ecu_name in ecu_col_index:
                    ecu_base = ecu_name.split("_")[0] if '_' in ecu_name else ecu_name
                    domain_folder_name = get_domain_folder_name(ecu_base, domain_short)
                    ecu_folder_name = get_ecu_folder_name(domain_folder_name, ecu_base)
                    
                    if ecu_folder_name:
                        date_str = datetime.now().strftime("%d%m%Y")
                        output_ecu_filename = f"ATOM_CAN_Matrix_{ecu_versions[ecu_name]}_{date_str}_{ecu_name}"
                        dbc_path = f"{create_directory.creator.PATH_DOC}\\{domain_folder_name}\\{ecu_folder_name}\\{output_ecu_filename}.dbc"
                        
                        # if os.path.exists(dbc_path):
                        #     with open(dbc_path, "rb") as f:
                        #         st.download_button(
                        #             label=f"Download DBC for {ecu_name}",
                        #             data=f,
                        #             file_name=f"{output_ecu_filename}.dbc",
                        #             mime="application/octet-stream"
                                # )
            except Exception as e:
                progress_bar.empty()
                status_text.empty()
                st.error(f"🕭 Error processing file: {e}")